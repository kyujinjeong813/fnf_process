# https://github.com/huaying/instagram-crawler/blob/e252cf2e892057fd338d4ce1affa9dc76f1e813e/inscrawler/crawler.py
# https://www.popit.kr/web-scraping-by-selenium/ (속도개선을 위해 참고합스다)

from inscrawler import crawler
from datetime import date


today = date.today() # 시간까진 필요 없겠지...?
logging = crawler.Logging()
Crawler = crawler.InsCrawler(logging)

user_id = ['96._.bo','bumki___','dykeee','k1mmg',
'kirnhh_','hyeon__e_','hyuk222','cd.ps','hyungmiin','seop_song',
'o62oo','orada5','yooncarp','leewoon19','smxmu','drleejd_10','leetaekk','h___ryan','scott.im',
'hbro_ys','2seu1','im_pinggggg','gloyk91']

# get_user_posts 내의 함수를 get_post_key로 하면 최근 12개 글을 가져옴
# get_post_key2로 하면 최근 기준 글 max 35개 가져옴 (가끔 7개 이렇게 되는 거 왜그런지 확인할 것)
# 댓글 수집 추가 (전체 수집은 비효율 아닌가 / 미션글에 대해서만? / 어떤 형태로 저장할지?)
posts = []
for user in user_id:
    dict_user = {}
    dict_user['user'] = user
#    dict_user['user'] = Crawler.get_user_profile(user)
    key_sets = Crawler.get_user_posts(user)
    dict_posts = []
    for key in key_sets:
        dict_post = Crawler.get_details(key)
        dict_post["url"] = key
        dict_posts.append(dict_post)
    dict_user['contents'] = dict_posts
    posts.append(dict_user)

print(posts)

# 엑셀 파일로 저장
import openpyxl

workbook = openpyxl.Workbook()
sheet = workbook.active

row = 1
for i in range(len(posts)):
    num = len(posts[i]['contents'])
    cnt = 1
    while cnt < num:
        try:
            sheet.cell(row=row, column=1, value=posts[i]['user'])
            # 1번 로우, 2번 컬럼 key값

            dic = posts[i]['contents'][cnt]
            if 'key' in dic.keys():
                sheet.cell(row=row, column=2, value=dic['key'])
            if 'datetime' in dic.keys():
                sheet.cell(row=row, column=3, value=dic['datetime'])
            if 'like' in dic.keys():
                sheet.cell(row=row, column=4, value=dic['like'])
            if 'view' in dic.keys():
                sheet.cell(row=row, column=5, value=dic['view'])
            if 'hashtags' in dic.keys():
                tags = '#'.join(dic['hashtags'])
                sheet.cell(row=row, column=6, value=tags)
            # 순서는 필요하면 변경할 것
            sheet.cell(row=row, column=7, value=today)

            cnt += 1
            row += 1

        except IndexError:
            pass

workbook.save(filename="insta_crawling.xlsx")